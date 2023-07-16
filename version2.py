import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QVBoxLayout, QWidget,QPushButton,QFileDialog,QLineEdit,QHBoxLayout
from PyQt5.QtGui import QPixmap, QFont,QIcon
from PyQt5.QtCore import Qt
import pandas as pd
import xlsxwriter
import os


import zipfile
from io import BytesIO
import barcode
from barcode.writer import ImageWriter
from PIL import Image
import cv2
import numpy as np
import matplotlib.pyplot as plt
import keras_ocr
import math
from barcode.writer import SVGWriter




from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QFileDialog
import shutil


def downloadFile(Input,path):
    if Input=="text":
        
        
        
        # Specify the path and filename of the file to be downloaded
        

        # Prompt the user to select the download location
        save_path, _ = QFileDialog.getSaveFileName(None, "Save File", "", "Png (*.png)")

        if save_path:
            try:
            
                
                shutil.copy(path, save_path)
                print("File downloaded successfully.")
                os.remove(path)
            except Exception as e:
                print("Error while downloading file:", str(e))
        else:
            print("Download canceled.")
            
    elif Input=="upload":
        save_path, _ = QFileDialog.getSaveFileName(None, "Save File", "", "Excel files (*.xlsx)")

        if save_path:
            try:
                
                shutil.copy(path, save_path)
                print("File downloaded successfully.")
                os.remove(path)
            except Exception as e:
                print("Error while downloading file:", str(e))
        else:
            print("Download canceled.")
            
        
def generate_barcode(text, barcode_type='code39'):
    text=str(text)
    text=text.replace('\n', '')
    
    text=text.replace(' ', '')
    text=text.replace('NaN','')
    '''
    # Create a barcode object
    barcode_class = barcode.get_barcode_class(barcode_type)
    barcode_object = barcode_class(text, writer=ImageWriter,add_checksum=False)'''
    
    barcode_object=barcode.Code39(text)
    
    barcode_object.writer=ImageWriter()
    image=barcode_object.render(text='')
    filename = f'barcode_{text}'
    image.save(filename+'.png')
    
    
    

    # Set the filename for the barcode image
    
    
    print(text)
    '''
    filename = f'barcode_{text}'
    barcode_object.save(filename)
    barcode_object.save('text')
    
    
    
    
    
    
    image = cv2.imread(filename+".png")
    def midpoint(x1, y1, x2, y2):
        x_mid = int((x1 + x2)/2)
        y_mid = int((y1 + y2)/2)
        return (x_mid, y_mid)

    #Main function that detects text and inpaints. 
    #Inputs are the image path and kreas_ocr pipeline
    def inpaint_text(img_path, pipeline):
        # read the image 
        img = keras_ocr.tools.read(img_path) 
        
        # Recogize text (and corresponding regions)
        # Each list of predictions in prediction_groups is a list of
        # (word, box) tuples. 
        prediction_groups = pipeline.recognize([img])
        
        #Define the mask for inpainting
        mask = np.zeros(img.shape[:2], dtype="uint8")
        for box in prediction_groups[0]:
            x0, y0 = box[1][0]
            x1, y1 = box[1][1] 
            x2, y2 = box[1][2]
            x3, y3 = box[1][3] 
            
            x_mid0, y_mid0 = midpoint(x1, y1, x2, y2)
            x_mid1, y_mi1 = midpoint(x0, y0, x3, y3)
            
            #For the line thickness, we will calculate the length of the line between 
            #the top-left corner and the bottom-left corner.
            thickness = int(math.sqrt( (x2 - x1)**2 + (y2 - y1)**2 ))
            
            #Define the line and inpaint
            cv2.line(mask, (x_mid0, y_mid0), (x_mid1, y_mi1), 255,    
            thickness)
            inpainted_img = cv2.inpaint(img, mask, 7, cv2.INPAINT_NS)
                    
        return(inpainted_img)

    # keras-ocr will automatically download pretrained
    # weights for the detector and recognizer.
    pipeline = keras_ocr.pipeline.Pipeline()

    img_text_removed = inpaint_text(filename+'.png', pipeline)

    plt.imshow(img_text_removed)

    cv2.imwrite(filename+'.png', cv2.cvtColor(img_text_removed, cv2.COLOR_BGR2RGB))
    image=Image.open(filename+'.png')
    width = 150
    height = 150

    # Resize the image
    image = image.resize((width, height))
    image.save(filename+".png")'''
    return filename+".png"


if getattr(sys, 'frozen', False):
    # If the application is run as a bundle (e.g., PyInstaller executable)
    logo_path = sys._MEIPASS + '/logo.png'  # Update 'logo.png' to the name of your logo file
else:
    # If the application is run as a standalone script
    logo_path = 'logo.png' 
# Get the current directory
current_dir = os.path.dirname(os.path.abspath(__file__))

# Set the current directory
os.chdir(current_dir)

app = QApplication(sys.argv)

def window_1(app):
    

    # Create a QMainWindow
    window = QMainWindow()
    window.setWindowTitle("Bar Code Generator")

    window.setFixedSize(700, 400)
    window.setWindowIcon(QIcon('logo.ico'))

    # Create a QVBoxLayout to hold the QLabel and text label
    layout = QVBoxLayout()

    # Add the image QLabel
    image_label = QLabel()
    pixmap = QPixmap("logo.png")  # Replace with your image path
    image_label.setPixmap(pixmap)
    image_label.setAlignment(Qt.AlignCenter)
    layout.addWidget(image_label)

    # Add the text QLabel
    text_label = QLabel("Unlock Your Products' Potential with Seamless Bar Code Generation!")
    text_label.setAlignment(Qt.AlignCenter)
    text_label.setFont(QFont("Arial", 14))  # Increase font size (adjust as needed)
    text_label.setMargin(1)  # Adjust the margin around the text label
    layout.addWidget(text_label)
    
    
    # Create a QWidget to hold the layout
    central_widget = QWidget()
    central_widget.setLayout(layout)
    window.setCentralWidget(central_widget)

    #add button
    button_next=QPushButton("Next")
    button_close=QPushButton("Close")

    #add then to layout
    layout.addWidget(button_next)
    layout.addWidget(button_close)

    

    #Add color
    button_close.setStyleSheet("background-color: #436953;color:white;")
    button_next.setStyleSheet("background-color: #436953;color:white;")
    #ADD button functionality
    def button_function():
        sender = window.sender()
        if sender == button_close:
            window.close()
        elif sender == button_next:
            window.close()
            window_2(app)
    button_next.clicked.connect(button_function)
    button_close.clicked.connect(button_function)
    # Show the main window
    window.show()

    # Start the application event loop
    sys.exit(app.exec_())

def window_2(app):
    

    # Create a QMainWindow
    window = QMainWindow()
    window.setWindowTitle("Bar Code Generator")
    window.setFixedSize(700, 400)
    window.setWindowIcon(QIcon('logo.ico'))
    layout = QVBoxLayout()
    
    central_widget = QWidget()
    central_widget.setLayout(layout)
    window.setCentralWidget(central_widget)
    
    
    
    # Add a label for the instruction on Uploading
    
    
    
    
    
    text_entry = QLineEdit(window)
    #Add styles for all the widgets
    
    text_entry.setFixedHeight(100)
    layout.addWidget(text_entry)
    
    
    
    
    # Function for tgetting the text
    def get_text():
        sender=window.sender()
        if sender==submit_text:
        
            entered_text = text_entry.text()
            image_name=generate_barcode(entered_text)
            Input="text"
            downloadFile(Input,image_name)
            
            
    
    
    upload_button = QPushButton("Upload File", window)
    
    upload_button.setGeometry(250, 400, 200, 50)
    
    upload_button.setStyleSheet("background-color:#436953;color:white;")
    
    # Add submit button for text-based input
    
    submit_text=QPushButton("Generate",window)
    
    submit_text.setStyleSheet("background-color:#436953;color:white;")
    
    submit_text.setFixedWidth(200)
    layout.addWidget(submit_text)
    
    submit_text.clicked.connect(lambda:get_text())
    
    
    

    
    # Set width for the QPushButton widget
    upload_button.setFixedWidth(200)
    # Create a QVBoxLayout to hold the QLabel and text label
    
    layout.addWidget(upload_button)
    
        
    
    
    
    
    
    def zip_files(file_paths, zip_name):
        with zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file in file_paths:
                zipf.write(file,os.path.basename(file))
    def upload():
        sender=window.sender()
        if sender==upload_button:
            
            file_dialog = QFileDialog()
            file_path, _ = file_dialog.getOpenFileName(window, "Upload File")
            if file_path:
                # Process the selected file
                

                
                data = pd.read_excel(file_path)
                image_list = []

                # Generate images and store their paths
                for index, row in data.iterrows():
                    
                    
                    image_name = generate_barcode(str(row.to_string(index=False)))
                    image_list.append(image_name)

                # Create the result workbook and worksheet
                workbook = xlsxwriter.Workbook("result.xlsx")
                worksheet = workbook.add_worksheet()

                # Iterate over the images and insert them into the worksheet
                for i, image_name in enumerate(image_list):
                    filename = image_name

                    file = open(filename, 'rb')
                    data_ = BytesIO(file.read())
                    file.close()

                    # Adjust the cell reference based on the current iteration
                    cell_reference = f'A{2 + i}'  # A1, A2, A3, ...

                    worksheet.insert_image(cell_reference, filename, {'image_data': data_})
                    
                #remove the images now
                for c in image_list:
                    os.remove(c)

                
                workbook.close()
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows

            file_path = 'result.xlsx'
            wb = openpyxl.load_workbook(file_path)

            # Step 2: Select the worksheet where you want to insert the data
            sheet_name = 'Sheet1'  # Replace 'Sheet1' with the actual sheet name
            ws = wb[sheet_name]

            # Step 3: Determine the starting cell for inserting the DataFrame
            start_row = ws.max_row + 1  # Start from the next available row

            # Step 4: Convert the DataFrame to rows
            rows = list(dataframe_to_rows(data, index=False, header=True))

            # Step 5: Insert the rows into the worksheet
            for row in rows:
                ws.append(row)

            # Step 6: Save the modified Excel file
            wb.save(file_path)
            
            filepath="result.xlsx"
            Input="upload"
            
            downloadFile(Input,filepath)

                        
        '''for i in data.barcode:
                            print(i)
                            image_name=generate_barcode(i)
                            image_fullname=image_name+".png"
                            list_images.append(image_fullname)
                            print(image_fullname)
                            zip_files(list_images, 'barcodes.zip')
                            
                            
                        for image in list_images:
                            os.remove(image)   
                        layout.addWidget(download_output)   '''

                            # Calculate the desired width and height for the image in the Word document
                    
                

    upload_button.clicked.connect(upload)
    
    
    window.show()
    


window_1(app)
